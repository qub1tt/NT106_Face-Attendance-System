import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from PyQt6 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook

# Khởi tạo Firebase với tệp JSON chứa khóa xác thực
cred = credentials.Certificate("ServiceAccountKey.json")

firebase_admin.initialize_app(cred, {
    "databaseURL":"https://faceregconition-80c55-default-rtdb.firebaseio.com/",
    "storageBucket":"faceregconition-80c55.appspot.com"
})

# Lấy reference đến nút "Students" trong Firebase Realtime Database
class_ref = db.reference('Classes')

# Lấy dữ liệu từ Firebase
class_data = class_ref.get()

# Lấy reference đến nút "Students" trong Firebase Realtime Database
students_ref = db.reference('Students')

# Lấy dữ liệu từ Firebase
s1 = students_ref.get()
student_data = {}

class Ui_StudentManagement(object):
    def setupUi(self, StudentManagement):
        StudentManagement.setObjectName("StudentManagement")
        StudentManagement.resize(1272, 680)
        StudentManagement.setStyleSheet("#StudentManagement{\n"
"    background-color: rgb(255,255,255);\n"
"}\n"
"\n"
"#Header {\n"
"    background-color: rgb(165, 213, 255);\n"
"}\n"
"\n"
"#Header #Logo{\n"
"    image: url(:/Pic/logo.png);\n"
"    border: none;\n"
"}\n"
"\n"
"#Header #NameSW{\n"
"    font-family: \"Robotol\", sans-serif;\n"
"    font-size: 25px;\n"
"    font-weight: bold;\n"
"}\n"
"\n"
"#result_frame {\n"
"    border-radius: 10px;\n"
"    border: 1px solid black;\n"
"    background-color: #fff;\n"
"}\n"
"\n"
"QTableWidget {\n"
"    border-radius: 3px;\n"
"    border: 1px solid #f0f0f0;\n"
"}\n"
"\n"
"QHeaderView::section {\n"
"    border: none;\n"
"    border-bottom: 1px solid black;\n"
"    padding: 3px 5px;\n"
"}\n"
"\n"
"QTableWidget::Item {\n"
"    border-bottom: 1px solid rgb(212, 212, 212);\n"
"    color: #000;\n"
"    padding-left: 3px;\n"
"}\n"
"\n"
"#btn_frame {\n"
"    border: 1px solid black;\n"
"    border-radius: 10px;\n"
"    background-color: rgb(255,255,255);\n"
"}\n"
"\n"
"#btn_frame QPushButton{\n"
"    background-color: #a5d5ff;\n"
"    border-radius: 10px;\n"
"}\n"
"\n"
"#btn_frame #delete_btn{\n"
"    background-color: rgb(255, 111, 111);\n"
"}\n"
"\n"
"#btn_frame #export_btn{\n"
"    background-color: rgb(74, 222, 128);\n"
"}\n"
"\n"
"#btn_frame QPushButton:hover{\n"
"    background-color: rgb(3, 105, 161); /* Màu nền mới khi hover */\n"
"    border-color: rgb(65, 173, 255);\n"
"    color: rgb(255, 255, 255);\n"
"}\n"
"#btn_frame #delete_btn:hover{\n"
"    background-color: rgb(185, 28, 28); /* Màu nền mới khi hover */\n"
"    border-color: rgb(65, 173, 255);\n"
"    color: rgb(255, 255, 255);\n"
"}\n"
"#btn_frame #export_btn:hover{\n"
"    background-color: rgb(185, 28, 28);\n"
"    color: rgb(255, 255, 255);\n"
"\n"
"\n"
"\n"
"\n"
"\n"
"")
        self.centralwidget = QtWidgets.QWidget(parent=StudentManagement)
        self.centralwidget.setObjectName("centralwidget")
        self.Header = QtWidgets.QFrame(parent=self.centralwidget)
        self.Header.setGeometry(QtCore.QRect(0, 0, 1281, 71))
        self.Header.setStyleSheet("")
        self.Header.setObjectName("Header")
        self.NameSW = QtWidgets.QLabel(parent=self.Header)
        self.NameSW.setGeometry(QtCore.QRect(30, 20, 501, 31))
        font = QtGui.QFont()
        font.setFamily("Robotol")
        font.setPointSize(-1)
        font.setBold(True)
        font.setWeight(75)
        self.NameSW.setFont(font)
        self.NameSW.setStyleSheet("")
        self.NameSW.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignLeft)
        self.NameSW.setObjectName("NameSW")
        self.label_class = QtWidgets.QLabel(parent=self.Header)
        self.label_class.setGeometry(QtCore.QRect(680, 20, 311, 31))
        font = QtGui.QFont()
        font.setFamily("Robotol")
        font.setBold(True)
        font.setWeight(75)
        self.label_class.setFont(font)
        self.label_class.setStyleSheet("")
        self.label_class.setText("")
        self.label_class.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignLeft)
        self.label_class.setObjectName("label_class")
        self.widget = QtWidgets.QWidget(parent=self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(0, 70, 1281, 631))
        self.widget.setStyleSheet("")
        self.widget.setObjectName("widget")
        self.result_frame = QtWidgets.QFrame(parent=self.widget)
        self.result_frame.setGeometry(QtCore.QRect(30, 120, 1211, 481))
        self.result_frame.setStyleSheet("")
        self.result_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.result_frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.result_frame.setObjectName("result_frame")
        self.gridLayout = QtWidgets.QGridLayout(self.result_frame)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(parent=self.result_frame)
        self.tableWidget.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        self.tableWidget.setStyleSheet("")
        self.tableWidget.setShowGrid(False)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(7)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(6, item)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(180)
        self.tableWidget.horizontalHeader().setSortIndicatorShown(False)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setColumnWidth(0, 100)
        self.tableWidget.setColumnWidth(1, 180)
        self.tableWidget.setColumnWidth(2, 180)
        self.tableWidget.setColumnWidth(3, 250)
        self.tableWidget.setColumnWidth(4, 250)
        self.tableWidget.setColumnWidth(5, 100)
        self.tableWidget.setColumnWidth(6, 100)
        self.gridLayout.addWidget(self.tableWidget, 0, 0, 1, 1)
        self.btn_frame = QtWidgets.QFrame(parent=self.widget)
        self.btn_frame.setGeometry(QtCore.QRect(30, 20, 1211, 80))
        self.btn_frame.setStyleSheet("")
        self.btn_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.btn_frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.btn_frame.setObjectName("btn_frame")
        self.update_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.update_btn.setGeometry(QtCore.QRect(550, 20, 111, 41))
        self.update_btn.setStyleSheet("")
        self.update_btn.setObjectName("update_btn")
        self.search_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.search_btn.setGeometry(QtCore.QRect(380, 20, 111, 41))
        self.search_btn.setStyleSheet("")
        self.search_btn.setObjectName("search_btn")
        self.delete_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.delete_btn.setGeometry(QtCore.QRect(890, 20, 111, 41))
        self.delete_btn.setStyleSheet("")
        self.delete_btn.setObjectName("delete_btn")
        self.read_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.read_btn.setGeometry(QtCore.QRect(210, 20, 111, 41))
        self.read_btn.setStyleSheet("")
        self.read_btn.setObjectName("read_btn")
        self.Diem_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.Diem_btn.setGeometry(QtCore.QRect(720, 20, 111, 41))
        self.Diem_btn.setStyleSheet("")
        self.Diem_btn.setObjectName("Diem_btn")
        self.export_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.export_btn.setGeometry(QtCore.QRect(1060, 20, 111, 41))
        self.export_btn.setStyleSheet("")
        self.export_btn.setObjectName("export_btn")
        self.class_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.class_btn.setGeometry(QtCore.QRect(40, 20, 111, 41))
        self.class_btn.setStyleSheet("")
        self.class_btn.setObjectName("class_btn")
        StudentManagement.setCentralWidget(self.centralwidget)

        self.retranslateUi(StudentManagement)
        QtCore.QMetaObject.connectSlotsByName(StudentManagement)
        StudentManagement.resizeEvent = self.on_window_resized

    def retranslateUi(self, StudentManagement):
        _translate = QtCore.QCoreApplication.translate
        StudentManagement.setWindowTitle(_translate("StudentManagement", "MainWindow"))
        self.NameSW.setText(_translate("StudentManagement", "STUDENT INFORMATION MANAGEMENT"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("StudentManagement", "StudentID"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("StudentManagement", "Name"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("StudentManagement", "Email"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("StudentManagement", "Faculty"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("StudentManagement", "Major"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("StudentManagement", "Year"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("StudentManagement", "Mark"))
        self.update_btn.setText(_translate("StudentManagement", "Update"))
        self.search_btn.setText(_translate("StudentManagement", "Search"))
        self.delete_btn.setText(_translate("StudentManagement", "Delete"))
        self.read_btn.setText(_translate("StudentManagement", "Read"))
        self.Diem_btn.setText(_translate("StudentManagement", "Mark"))
        self.export_btn.setText(_translate("StudentManagement", "Export"))
        self.class_btn.setText(_translate("StudentManagement", "Select Class"))

        self.class_btn.clicked.connect(self.select_class)
        self.read_btn.clicked.connect(self.load_data)
        self.search_btn.clicked.connect(self.search_data)
        self.update_btn.clicked.connect(self.update_data)
        self.delete_btn.clicked.connect(self.delete_data)
        self.Diem_btn.clicked.connect(self.calculate_data)
        self.export_btn.clicked.connect(self.export_data)
        
    def on_window_resized(self, event):
        # Lấy kích thước mới của cửa sổ
        window_size = event.size()

        # Cập nhật kích thước và vị trí của các thành phần
        self.centralwidget.setGeometry(0, 0, window_size.width(), window_size.height())
        self.Header.setGeometry(0, 0, window_size.width(), 71)
        self.widget.setGeometry(0, 71, window_size.width(), window_size.height() - 71)
        self.result_frame.setGeometry(30, 120, window_size.width() - 60, window_size.height() - 191)
        self.btn_frame.setGeometry(30, 20, window_size.width() - 60, 80)

    def select_class(self):
        if class_data:
            class_names = list(class_data.keys())
            selected_class, ok =  QtWidgets.QInputDialog.getItem(None, "Select Class", "Choose a class:", class_names, 0, False)
            if ok and selected_class:
                self.label_class.setText("Class: " + selected_class)
                font = QtGui.QFont()
                font.setPointSize(20)
                font.setBold(True)
                self.label_class.setFont(font)
                # Filter student data based on selected class
                self.tableWidget.clearContents()
                self.tableWidget.setRowCount(0)
                self.filter_student_data(selected_class)
        else:
            QtWidgets.QMessageBox.warning(None, "No Classes", "No classes found, please try again later.")

    def filter_student_data(self, selected_class):
        student_data.clear()
        if s1:
            for student_id, student_info in s1.items():
                classes = student_info.get("Classes", {})
                if selected_class in classes:
                    student_data[student_id] = student_info
        self.load_data()
        

    def load_data(self):
        if student_data:
            self.tableWidget.setRowCount(len(student_data))
            row = 0
            for student_id, student_info in student_data.items():
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(student_id))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(student_info.get("Name", "")))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(student_info.get("Email", "")))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(student_info.get("Faculty", "")))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(student_info.get("Major", "")))
                year_str = str(student_info.get("Year", ""))
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(year_str))
                
                row += 1
            set_read_only_flags(self.tableWidget)
        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")
    
    # def search_data(self): 
    # # Phương thức để tìm kiếm sinh viên
    #     if student_data:
    #         text, ok = QtWidgets.QInputDialog.getText(None, 'Search Student', 'Enter Student ID:')
    #         if ok:
    #             if text.strip():  # Kiểm tra xem người dùng đã nhập dữ liệu hay không
    #                 student_info = student_data.get(text)
    #                 if student_info:
    #                     self.tableWidget.setRowCount(1)
    #                     self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem(text))
    #                     self.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(student_info.get("Name", "")))
    #                     self.tableWidget.setItem(0, 2, QtWidgets.QTableWidgetItem(student_info.get("Email", "")))
    #                     self.tableWidget.setItem(0, 3, QtWidgets.QTableWidgetItem(student_info.get("Faculty", "")))
    #                     self.tableWidget.setItem(0, 4, QtWidgets.QTableWidgetItem(student_info.get("Major", "")))
    #                     year_str = str(student_info.get("Year", ""))
    #                     self.tableWidget.setItem(0, 5, QtWidgets.QTableWidgetItem(year_str))
    #                     set_read_only_flags(self.tableWidget)
    #                 else:
    #                     QtWidgets.QMessageBox.warning(None, 'Search Student', 'Student ID not found.')
    #             else:
    #                 QtWidgets.QMessageBox.warning(None, 'Search Student', 'Please enter a Student ID.')
    #     else:
    #         QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")

    def search_data(self):
    # Phương thức để tìm kiếm sinh viên
        if student_data:
            # Tạo danh sách các trường dữ liệu để người dùng chọn
            fields = ["Student ID", "Name", "Email", "Faculty", "Major", "Year"]
            field, ok = QtWidgets.QInputDialog.getItem(None, "Select Field to Search", "Field:", fields, 0, False)
            if ok:
                text, ok = QtWidgets.QInputDialog.getText(None, f'Search Student by {field}', f'Enter {field}:')
                if ok:
                    if text.strip():  # Kiểm tra xem người dùng đã nhập dữ liệu hay không
                        # Thực hiện tìm kiếm dựa trên trường dữ liệu được chọn
                        if field == "Student ID":
                            student_info = student_data.get(text)
                            if student_info:
                                self.display_search_result(text, student_info)
                            else:
                                QtWidgets.QMessageBox.warning(None, 'Search Student', 'Student ID not found.')
                        else:
                            # Tìm kiếm trong tất cả sinh viên và hiển thị kết quả phù hợp
                            found_students = []
                            for student_id, student_info in student_data.items():
                                if student_info.get(field, "") == text:
                                    found_students.append((student_id, student_info))
                            if found_students:
                                for student_id, student_info in found_students:
                                    self.display_search_result(student_id, student_info)
                            else:
                                QtWidgets.QMessageBox.warning(None, 'Search Student', 'Cannot find, Please try again.')
                    else:
                        QtWidgets.QMessageBox.warning(None, 'Search Student', f'Please enter a {field}.')
        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")

    def display_search_result(self, student_id, student_info):
        # Hiển thị kết quả tìm kiếm trên tableWidget
        self.tableWidget.setRowCount(1)
        self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem(student_id))
        self.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(student_info.get("Name", "")))
        self.tableWidget.setItem(0, 2, QtWidgets.QTableWidgetItem(student_info.get("Email", "")))
        self.tableWidget.setItem(0, 3, QtWidgets.QTableWidgetItem(student_info.get("Faculty", "")))
        self.tableWidget.setItem(0, 4, QtWidgets.QTableWidgetItem(student_info.get("Major", "")))
        year_str = str(student_info.get("Year", ""))
        self.tableWidget.setItem(0, 5, QtWidgets.QTableWidgetItem(year_str))
        set_read_only_flags(self.tableWidget)
    

    def update_data(self):
        if student_data:
            text, ok = QtWidgets.QInputDialog.getText(None, 'Update Student', 'Enter Student ID:')
            if ok:
                student_info = student_data.get(text)
                if student_info:
                    items = ["Name", "Email", "Faculty", "Major", "Year"]
                    item, ok = QtWidgets.QInputDialog.getItem(None, "Select Item to Update", "Item:", items, 0, False)
                    if ok and item:
                        new_value, ok = QtWidgets.QInputDialog.getText(None, 'Update Student', f'Enter new {item}:')
                        if ok:
                            student_info[item] = new_value
                            students_ref.child(text).set(student_info)
                            QtWidgets.QMessageBox.information(None, 'Update Student', 'Student updated successfully.')
                            self.load_data()
                else:
                    QtWidgets.QMessageBox.warning(None, 'Update Student', 'Student ID not found.')
        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")
                
    def delete_data(self):
        if student_data:
            text, ok = QtWidgets.QInputDialog.getText(None, 'Delete Student', 'Enter Student ID:')
            if ok:
                student_info = student_data.get(text)
                if student_info:
                    confirm = QtWidgets.QMessageBox.question(None, 'Delete Student', 'Are you sure you want to delete this student?', QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No)
                    if confirm == QtWidgets.QMessageBox.StandardButton.Yes:
                        # Xóa dữ liệu từ Firebase
                        students_ref.child(text).delete()
                        QtWidgets.QMessageBox.information(None, 'Delete Student', 'Student deleted successfully.')
                        
                        # Xóa dữ liệu từ student_data
                        del student_data[text]
                        
                        # Cập nhật lại dữ liệu trên Table Widget
                        self.load_data()
                else:
                    QtWidgets.QMessageBox.warning(None, 'Delete Student', 'Student ID not found.')
        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")
                
    # def calculate_data(self):
    #     if student_data:
    #         if self.tableWidget.rowCount() == 0:
    #             QtWidgets.QMessageBox.warning(None, 'Calculate', 'No data to calculate.')
    #             return

    #         # Lấy tên lớp đang được chọn
    #         selected_class_name = self.label_class.text().split(": ")[1]

    #         # Tạo một danh sách các sinh viên thuộc lớp đã chọn
    #         selected_class_students = [student_id for student_id, student_info in student_data.items() if selected_class_name in student_info.get("Classes", [])]

    #         for row, student_id in enumerate(selected_class_students):
    #             student_info = student_data.get(student_id)
    #             if student_info:
    #                 # Lấy giá trị AttendanceCount từ dữ liệu của sinh viên
    #                 attendance_count = student_info.get("Classes", {}).get(selected_class_name, {}).get("AttendanceCount", 0)

    #                 # Tính điểm dựa trên AttendanceCount (ví dụ: 1 điểm cho mỗi lần điểm danh)
    #                 # Bạn có thể điều chỉnh công thức tính điểm theo nhu cầu của bạn
    #                 points = attendance_count
                    
    #                 # Tạo một QTableWidgetItem để hiển thị điểm
    #                 points_item = QtWidgets.QTableWidgetItem(str(points))
    #                 # Đặt điểm vào cột "Điểm" của hàng hiện tại
    #                 self.tableWidget.setItem(row, 6, points_item)
    #                 set_read_only_flags(self.tableWidget)

    #         # Hiển thị thông báo khi tính toán hoàn thành
    #         QtWidgets.QMessageBox.information(None, 'Calculate', 'Calculation completed successfully.')

    #     else:
    #         QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")

    def calculate_data(self):
        if student_data:
            if self.tableWidget.rowCount() == 0:
                QtWidgets.QMessageBox.warning(None, 'Calculate', 'No data to calculate.')
                return

            # Lấy tên lớp đang được chọn
            selected_class_name = self.label_class.text().split(": ")[1]

            # Tạo một danh sách các sinh viên thuộc lớp đã chọn
            selected_class_students = [student_id for student_id, student_info in student_data.items() if selected_class_name in student_info.get("Classes", [])]

            # Hiển thị hộp thoại cho người dùng để chọn liệu muốn sắp xếp điểm hay không
            reply = QtWidgets.QMessageBox.question(None, 'Sort Points', 'Do you want to sort points?', QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No)
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                # Hiển thị hộp thoại cho người dùng để chọn kiểu sắp xếp
                sort_options = ['Ascending', 'Descending']
                sort, ok = QtWidgets.QInputDialog.getItem(None, 'Sort Options', 'Choose sort order:', sort_options, 0, False)
                if ok and sort:
                    reverse = True if sort == 'Descending' else False
                    selected_class_students.sort(key=lambda student_id: student_data[student_id]['Classes'][selected_class_name]['AttendanceCount'], reverse=reverse)

            for row, student_id in enumerate(selected_class_students):
                student_info = student_data.get(student_id)
                if student_info:
                    # Lấy giá trị AttendanceCount từ dữ liệu của sinh viên
                    attendance_count = student_info.get("Classes", {}).get(selected_class_name, {}).get("AttendanceCount", 0)

                    # Tính điểm dựa trên AttendanceCount (ví dụ: 1 điểm cho mỗi lần điểm danh)
                    # Bạn có thể điều chỉnh công thức tính điểm theo nhu cầu của bạn
                    points = attendance_count
                    
                    # Tạo một QTableWidgetItem để hiển thị điểm
                    points_item = QtWidgets.QTableWidgetItem(str(points))
                    # Đặt điểm vào cột "Điểm" của hàng hiện tại
                    self.tableWidget.setItem(row, 6, points_item)
                    set_read_only_flags(self.tableWidget)

            # Hiển thị thông báo khi tính toán hoàn thành
            QtWidgets.QMessageBox.information(None, 'Calculate', 'Calculation completed successfully.')

        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")



    def export_data(self):
        if student_data:
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Save File", "", "Excel Files (*.xlsx)")
            if file_path:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Student Data"

                # Ghi tiêu đề cột
                for column in range(self.tableWidget.columnCount()):
                    sheet.cell(row=1, column=column + 1, value=self.tableWidget.horizontalHeaderItem(column).text())

                # Ghi dữ liệu từ tableWidget vào file Excel
                for row in range(self.tableWidget.rowCount()):
                    for column in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, column)
                        if item is not None:
                            sheet.cell(row=row + 2, column=column + 1, value=item.text())

                try:
                    workbook.save(file_path)
                    QtWidgets.QMessageBox.information(None, 'Export', f'Data exported to {file_path} successfully.')
                except Exception as e:
                    QtWidgets.QMessageBox.warning(None, 'Export', f'Error occurred while exporting data: {str(e)}')
        else:
            QtWidgets.QMessageBox.warning(None, "Warning", "Please select a class.")
            
def set_read_only_flags(table_widget):
    for row in range(table_widget.rowCount()):
        for column in range(table_widget.columnCount()):
            item = table_widget.item(row, column)
            if item:
                item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    StudentManagement = QtWidgets.QMainWindow()
    ui = Ui_StudentManagement()
    ui.setupUi(StudentManagement)
    StudentManagement.show()
    sys.exit(app.exec())