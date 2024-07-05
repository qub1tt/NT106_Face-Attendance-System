import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from PyQt6 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
import requests

from firebaseconfig import class_ref, students_ref

class_data = class_ref.get()
s1 = students_ref.get()
student_data = {}


class SearchDialog(QtWidgets.QDialog):
    def __init__(self, class_names):
        super().__init__()

        # Lưu danh sách tên lớp vào thuộc tính của lớp
        self.class_names = class_names

        # Thiết lập tiêu đề của cửa sổ
        self.setWindowTitle("Select Class")
        
        self.setFixedSize(445, 225)

        # Tạo layout chính của cửa sổ
        main_layout = QtWidgets.QVBoxLayout()

        # Layout chứa ô nhập và combo box
        search_layout = QtWidgets.QHBoxLayout()
        self.search_input = QtWidgets.QLineEdit()  # Ô nhập văn bản
        self.search_input.setFixedSize(250, 35)

        # Tạo Completer và đặt cài đặt cho Completer
        self.completer = QtWidgets.QCompleter(class_names)
        self.completer.setCaseSensitivity(QtCore.Qt.CaseSensitivity.CaseInsensitive)  # Đảm bảo không phân biệt chữ hoa, chữ thường
        self.search_input.setCompleter(self.completer)

        self.search_input.setPlaceholderText("Enter class name")  # Placeholder cho ô nhập

        search_layout.addWidget(self.search_input)

        self.class_combo = QtWidgets.QComboBox()  # Combo box
        self.class_combo.setFixedSize(100, 35)
        self.class_combo.addItem("Select Class")
        self.class_combo.addItems(class_names)  # Thêm các tên lớp vào combo box
        search_layout.addWidget(self.class_combo)

        # Thêm các layout vào layout chính
        main_layout.addLayout(search_layout)
        
        # Layout chứa nút OK
        button_layout = QtWidgets.QHBoxLayout()
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setFixedSize(100, 30)  # Đặt kích thước cố định cho nút OK
        button_layout.addWidget(self.ok_button)

        # Thiết lập layout chính cho cửa sổ
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)

        # Kết nối tín hiệu của combo box với hàm cập nhật ô nhập
        self.class_combo.currentIndexChanged.connect(self.update_search_input)

        # Thêm bộ lọc tùy chỉnh cho Completer
        self.completer.setFilterMode(QtCore.Qt.MatchFlag.MatchContains)
        self.search_input.textChanged.connect(self.filter_completion)
        
        self.setStyleSheet(
            "QPushButton{min-width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);} QLineEdit{width:250px; height: 35px;font-size: 15px; padding-left: 5px; border: 1px solid black; border-radius: 5px} QCompleter{font-size: 15px} QComboBox{padding-left: 5px;  height: 35px; font-size: 15px;}"
            )

    # Hàm cập nhật ô nhập khi chọn một lớp từ combo box
    def update_search_input(self, index):
        if index > 0:
            self.search_input.setText(self.class_combo.currentText())
        else:
            self.search_input.setText(None)


    # Hàm trả về lớp được chọn
    def get_selected_class(self):
        if self.search_input.text():
            return self.search_input.text()
        else:
            return self.class_combo.currentText()

    # Hàm lọc gợi ý dựa trên chuỗi nhập của người dùng
    def filter_completion(self, text):
        filtered_class_names = [name for name in self.class_names if text.lower() in name.lower()]
        model = QtCore.QStringListModel(filtered_class_names)
        self.completer.setModel(model)

class AdvancedSearchDialog(QtWidgets.QDialog):
    def __init__(self, student_ids):
        super().__init__()

        self.student_ids = student_ids
        self.st = []
        self.s8 = []

        # Tạo danh sách self.st chứa thông tin của từng sinh viên
        for student_id in student_ids:
            student_info = student_data.get(student_id, {})
            if student_info:
                info_str = f"{student_id}"
                self.st.append(info_str)

        self.setWindowTitle("Find Student ID")
        self.setFixedSize(445, 225)

        main_layout = QtWidgets.QVBoxLayout()

        search_layout = QtWidgets.QHBoxLayout()
        self.search_input = QtWidgets.QLineEdit()
        self.search_input.setFixedSize(250, 35)

        # Sử dụng danh sách self.st cho QCompleter
        self.completer = QtWidgets.QCompleter(self.st)
        self.completer.setCaseSensitivity(QtCore.Qt.CaseSensitivity.CaseInsensitive)
        self.search_input.setCompleter(self.completer)

        self.search_input.setPlaceholderText("Enter Student ID")
        search_layout.addWidget(self.search_input)

        main_layout.addLayout(search_layout)

        button_layout = QtWidgets.QHBoxLayout()
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setFixedSize(100, 30)
        button_layout.addWidget(self.ok_button)

        self.setStyleSheet(
        "QPushButton{min-width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);} QLineEdit{width:250px; height: 35px;font-size: 15px; padding-left: 5px; border: 1px solid black; border-radius: 5px} QCompleter{font-size: 15px} QComboBox{padding-left: 5px;  height: 35px; font-size: 15px;}"
        )

        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)
        self.completer.setFilterMode(QtCore.Qt.MatchFlag.MatchContains)
        self.search_input.textChanged.connect(self.filter_completion)

    def filter_completion(self, text):
        # Lọc và giới hạn ký tự
        filtered_student_info = []
        for i in range(len(self.st)):
            info = self.st[i]
            if text.lower() in info.lower():
                filtered_student_info.append(info)
        # Cập nhật model cho completer
        model = QtCore.QStringListModel(filtered_student_info)
        self.completer.setModel(model)

    def get_selected_student(self):
        return self.search_input.text()
    
class UpdateDialog(QtWidgets.QDialog):
    def __init__(self, field_names):
        super().__init__()

        # Lưu danh sách tên trường cần update vào thuộc tính của lớp
        self.field_names = field_names

        # Thiết lập tiêu đề của cửa sổ
        self.setWindowTitle("Update Student")
        self.setFixedSize(445, 225)

        # Tạo layout chính của cửa sổ
        main_layout = QtWidgets.QVBoxLayout()

        # Layout chứa ô combo box để chọn trường cần update
        field_layout = QtWidgets.QHBoxLayout()
        self.field_combo = QtWidgets.QComboBox()  # Combo box
        self.field_combo.setFixedSize(150, 35)
        self.field_combo.addItem("Select Field")
        self.field_combo.addItems(field_names)  # Thêm các tên trường vào combo box
        field_layout.addWidget(self.field_combo)

        # Thêm layout chọn trường vào layout chính
        main_layout.addLayout(field_layout)

        # Layout chứa ô nhập giá trị mới
        value_layout = QtWidgets.QHBoxLayout()
        self.value_input = QtWidgets.QLineEdit()  # Ô nhập văn bản
        self.value_input.setFixedSize(250, 35)
        self.value_input.setEnabled(False)  # Không cho phép nhập liệu cho đến khi chọn trường
        self.value_input.setText("Please Select Field")
        value_layout.addWidget(self.value_input)

        # Thêm layout nhập giá trị mới vào layout chính
        main_layout.addLayout(value_layout)

        # Layout chứa nút OK
        button_layout = QtWidgets.QHBoxLayout()
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setFixedSize(100, 30)  # Đặt kích thước cố định cho nút OK
        button_layout.addWidget(self.ok_button)

        # Thiết lập layout chính cho cửa sổ
        main_layout.addLayout(button_layout)
        self.setLayout(main_layout)

        # Kết nối tín hiệu của combo box với hàm cập nhật ô nhập
        self.field_combo.currentIndexChanged.connect(self.update_value_input)

        self.setStyleSheet(
            "QPushButton{min-width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);} QLineEdit{width:250px; height: 35px;font-size: 15px; padding-left: 5px; border: 1px solid black; border-radius: 5px} QComboBox{padding-left: 5px;  height: 35px; font-size: 15px;}"
        )

    # Hàm cập nhật ô nhập khi chọn một trường từ combo box
    def update_value_input(self, index):
        if index > 0:
            selected_field = self.field_combo.currentText()
            current_value = ""  # Khởi tạo giá trị mặc định là rỗng

            # Lấy giá trị hiện tại của trường dữ liệu từ dữ liệu sinh viên
            if selected_field in s1:
                current_value = s1[selected_field]

            # Hiển thị giá trị hiện tại của trường dữ liệu trong ô nhập
            self.value_input.setText(current_value)
            self.value_input.setEnabled(True)
        else:
            self.value_input.setText("Please Select Field")
            self.value_input.setEnabled(False)

    # Hàm trả về trường được chọn cần update và giá trị mới nhập
    def get_selected_field_and_value(self):
        selected_field = self.field_combo.currentText()
        value = self.value_input.text()
        return selected_field, value


class Ui_StudentManagement(object):
    def setupUi(self, StudentManagement):
        StudentManagement.setObjectName("StudentManagement")
        StudentManagement.resize(1272, 680)
        StudentManagement.setStyleSheet("#StudentManagement{\n"
"    background-color: rgb(255,255,255);\n"
"}\n"
"\n"
"#Header {\n"
"    background-color: rgb(165, 213, 255);\n"
"}\n"
"\n"
"#Header #Logo{\n"
"    image: url(:/Pic/logo.png);\n"
"    border: none;\n"
"}\n"
"\n"
"#Header #NameSW{\n"
"    font-family: \"Robotol\", sans-serif;\n"
"    font-size: 25px;\n"
"    font-weight: bold;\n"
"}\n"
"\n"
"#result_frame {\n"
"    border-radius: 10px;\n"
"    border: 1px solid black;\n"
"    background-color: #fff;\n"
"}\n"
"\n"
"QTableWidget {\n"
"    border-radius: 3px;\n"
"    border: 1px solid #f0f0f0;\n"
"}\n"
"\n"
"QHeaderView::section {\n"
"    border: none;\n"
"    border-bottom: 1px solid black;\n"
"    padding: 3px 5px;\n"
"}\n"
"\n"
"QTableWidget::Item {\n"
"    border-bottom: 1px solid rgb(212, 212, 212);\n"
"    color: #000;\n"
"    padding-left: 3px;\n"
"}\n"
"\n"
"#btn_frame {\n"
"    border: 1px solid black;\n"
"    border-radius: 10px;\n"
"    background-color: rgb(255,255,255);\n"
"}\n"
"\n"
"#btn_frame QPushButton{\n"
"    background-color: #a5d5ff;\n"
"    border-radius: 10px;\n"
"}\n"
"\n"
"#btn_frame #delete_btn{\n"
"    background-color: rgb(255, 111, 111);\n"
"}\n"
"\n"
"#btn_frame #export_btn{\n"
"    background-color: rgb(74, 222, 128);\n"
"}\n"
"\n"
"#btn_frame QPushButton:hover{\n"
"    background-color: rgb(3, 105, 161); /* Màu nền mới khi hover */\n"
"    border-color: rgb(65, 173, 255);\n"
"    color: rgb(255, 255, 255);\n"
"}\n"
"#btn_frame #delete_btn:hover{\n"
"    background-color: rgb(185, 28, 28); /* Màu nền mới khi hover */\n"
"    border-color: rgb(65, 173, 255);\n"
"    color: rgb(255, 255, 255);\n"
"}\n"
"#btn_frame #export_btn:hover{\n"
"    background-color: rgb(185, 28, 28);\n"
"    color: rgb(255, 255, 255);\n"
"\n"
"\n"
"\n"
"\n"
"\n"
"")
        self.centralwidget = QtWidgets.QWidget(parent=StudentManagement)
        self.centralwidget.setObjectName("centralwidget")
        self.Header = QtWidgets.QFrame(parent=self.centralwidget)
        self.Header.setGeometry(QtCore.QRect(0, 0, 1281, 71))
        self.Header.setStyleSheet("")
        self.Header.setObjectName("Header")
        self.NameSW = QtWidgets.QLabel(parent=self.Header)
        self.NameSW.setGeometry(QtCore.QRect(30, 20, 501, 31))
        font = QtGui.QFont()
        font.setFamily("Robotol")
        font.setPointSize(-1)
        font.setBold(True)
        font.setWeight(75)
        self.NameSW.setFont(font)
        self.NameSW.setStyleSheet("")
        self.NameSW.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignLeft)
        self.NameSW.setObjectName("NameSW")
        self.label_class = QtWidgets.QLabel(parent=self.Header)
        self.label_class.setGeometry(QtCore.QRect(680, 20, 311, 31))
        font = QtGui.QFont()
        font.setFamily("Robotol")
        font.setBold(True)
        font.setWeight(75)
        self.label_class.setFont(font)
        self.label_class.setStyleSheet("")
        self.label_class.setText("")
        self.label_class.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignLeft)
        self.label_class.setObjectName("label_class")
        self.widget = QtWidgets.QWidget(parent=self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(0, 70, 1281, 631))
        self.widget.setStyleSheet("")
        self.widget.setObjectName("widget")
        self.result_frame = QtWidgets.QFrame(parent=self.widget)
        self.result_frame.setGeometry(QtCore.QRect(30, 120, 1211, 481))
        self.result_frame.setStyleSheet("")
        self.result_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.result_frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.result_frame.setObjectName("result_frame")
        self.gridLayout = QtWidgets.QGridLayout(self.result_frame)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(parent=self.result_frame)
        self.tableWidget.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        self.tableWidget.setStyleSheet("")
        self.tableWidget.setShowGrid(False)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(7)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(6, item)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(180)
        self.tableWidget.horizontalHeader().setSortIndicatorShown(False)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setColumnWidth(0, 100)
        self.tableWidget.setColumnWidth(1, 180)
        self.tableWidget.setColumnWidth(2, 180)
        self.tableWidget.setColumnWidth(3, 250)
        self.tableWidget.setColumnWidth(4, 250)
        self.tableWidget.setColumnWidth(5, 100)
        self.tableWidget.setColumnWidth(6, 100)
        self.gridLayout.addWidget(self.tableWidget, 0, 0, 1, 1)
        self.btn_frame = QtWidgets.QFrame(parent=self.widget)
        self.btn_frame.setGeometry(QtCore.QRect(30, 20, 1211, 80))
        self.btn_frame.setStyleSheet("")
        self.btn_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.btn_frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.btn_frame.setObjectName("btn_frame")
        self.update_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.update_btn.setGeometry(QtCore.QRect(550, 20, 111, 41))
        self.update_btn.setStyleSheet("")
        self.update_btn.setObjectName("update_btn")
        self.search_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.search_btn.setGeometry(QtCore.QRect(380, 20, 111, 41))
        self.search_btn.setStyleSheet("")
        self.search_btn.setObjectName("search_btn")
        self.delete_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.delete_btn.setGeometry(QtCore.QRect(890, 20, 111, 41))
        self.delete_btn.setStyleSheet("")
        self.delete_btn.setObjectName("delete_btn")
        self.read_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.read_btn.setGeometry(QtCore.QRect(210, 20, 111, 41))
        self.read_btn.setStyleSheet("")
        self.read_btn.setObjectName("read_btn")
        self.Diem_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.Diem_btn.setGeometry(QtCore.QRect(720, 20, 111, 41))
        self.Diem_btn.setStyleSheet("")
        self.Diem_btn.setObjectName("Diem_btn")
        self.export_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.export_btn.setGeometry(QtCore.QRect(1060, 20, 111, 41))
        self.export_btn.setStyleSheet("")
        self.export_btn.setObjectName("export_btn")
        self.class_btn = QtWidgets.QPushButton(parent=self.btn_frame)
        self.class_btn.setGeometry(QtCore.QRect(40, 20, 111, 41))
        self.class_btn.setStyleSheet("")
        self.class_btn.setObjectName("class_btn")
        StudentManagement.setCentralWidget(self.centralwidget)

        self.retranslateUi(StudentManagement)
        QtCore.QMetaObject.connectSlotsByName(StudentManagement)
        StudentManagement.resizeEvent = self.on_window_resized

    def retranslateUi(self, StudentManagement):
        _translate = QtCore.QCoreApplication.translate
        StudentManagement.setWindowTitle(_translate("StudentManagement", "MainWindow"))
        self.NameSW.setText(_translate("StudentManagement", "STUDENT INFORMATION MANAGEMENT"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("StudentManagement", "StudentID"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("StudentManagement", "Name"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("StudentManagement", "Email"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("StudentManagement", "Faculty"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("StudentManagement", "Major"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("StudentManagement", "Year"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("StudentManagement", "Mark"))
        self.update_btn.setText(_translate("StudentManagement", "Update"))
        self.search_btn.setText(_translate("StudentManagement", "Search"))
        self.delete_btn.setText(_translate("StudentManagement", "Delete"))
        self.read_btn.setText(_translate("StudentManagement", "Read"))
        self.Diem_btn.setText(_translate("StudentManagement", "Mark"))
        self.export_btn.setText(_translate("StudentManagement", "Export"))
        self.class_btn.setText(_translate("StudentManagement", "Select Class"))

        self.class_btn.clicked.connect(self.select_class)
        self.read_btn.clicked.connect(self.load_data)
        self.search_btn.clicked.connect(self.search_data)
        self.update_btn.clicked.connect(self.update_data)
        self.delete_btn.clicked.connect(self.delete_data)
        self.Diem_btn.clicked.connect(self.calculate_data)
        self.export_btn.clicked.connect(self.export_data)
        
        
    def on_window_resized(self, event):
        # Lấy kích thước mới của cửa sổ
        window_size = event.size()

        # Cập nhật kích thước và vị trí của các thành phần
        self.centralwidget.setGeometry(0, 0, window_size.width(), window_size.height())
        self.Header.setGeometry(0, 0, window_size.width(), 71)
        self.widget.setGeometry(0, 71, window_size.width(), window_size.height() - 71)
        self.result_frame.setGeometry(30, 120, window_size.width() - 60, window_size.height() - 191)
        self.btn_frame.setGeometry(30, 20, window_size.width() - 60, 80)

    def select_class(self):
        if class_data:
            class_names = list(class_data.keys())
            # Hiển thị dialog tìm kiếm và chọn lớp
            search_dialog = SearchDialog(class_names)
            if search_dialog.exec():
                selected_class = search_dialog.get_selected_class()
                
                if selected_class:
                    if selected_class in class_names:
                        self.label_class.setText("Class: " + selected_class)
                        font = QtGui.QFont()
                        font.setPointSize(20)
                        font.setBold(True)
                        self.label_class.setFont(font)
                        # Lọc dữ liệu sinh viên dựa trên lớp đã chọn
                        self.tableWidget.clearContents()
                        self.tableWidget.setRowCount(0)
                        self.filter_student_data(selected_class)
                    else:
                        qmb_custom("No Classes", "No classes found, please try again later.")
        else:
            qmb_custom("No Classes", "No classes found, please try again later.")
            

    def filter_student_data(self, selected_class):
        student_data.clear()
        if s1:
            for student_id, student_info in s1.items():
                classes = student_info.get("Classes", {})
                if selected_class in classes:
                    student_data[student_id] = student_info
        self.load_data()
        

    def load_data(self):
        if student_data:
            self.tableWidget.setRowCount(len(student_data))
            row = 0
            for student_id, student_info in student_data.items():
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(student_id))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(student_info.get("Name", "")))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(student_info.get("Email", "")))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(student_info.get("Faculty", "")))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(student_info.get("Major", "")))
                year_str = str(student_info.get("Year", ""))
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(year_str))
                
                row += 1
            set_read_only_flags(self.tableWidget)
        else:
            qmb_custom("Warning", "Please select a class.")

    def search_data(self):
        if student_data:
            # Lấy danh sách các sinh viên
            students = list(student_data.keys())

            # Hiển thị cửa sổ tìm kiếm nâng cao và lấy ID của sinh viên được chọn
            search_dialog = AdvancedSearchDialog(students)
            if search_dialog.exec():
                selected_student = search_dialog.get_selected_student()
                if selected_student in student_data:
                    student_info = student_data[selected_student]
                    self.display_search_result(selected_student, student_info)
                elif selected_student == "":
                    qmb_custom('Search Student', 'Please enter a Student ID.')
                    self.search_data()
                else:
                    qmb_custom('Search Student', 'Student ID not found.')
                    self.search_data()
        else:
            qmb_custom("Warning", "Please select a class.")

    def display_search_result(self, student_id, student_info):
        # Hiển thị kết quả tìm kiếm trên tableWidget
        self.tableWidget.setRowCount(1)
        self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem(student_id))
        self.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(student_info.get("Name", "")))
        self.tableWidget.setItem(0, 2, QtWidgets.QTableWidgetItem(student_info.get("Email", "")))
        self.tableWidget.setItem(0, 3, QtWidgets.QTableWidgetItem(student_info.get("Faculty", "")))
        self.tableWidget.setItem(0, 4, QtWidgets.QTableWidgetItem(student_info.get("Major", "")))
        year_str = str(student_info.get("Year", ""))
        self.tableWidget.setItem(0, 5, QtWidgets.QTableWidgetItem(year_str))
        set_read_only_flags(self.tableWidget)
    

    def update_data(self):
        if student_data:
            # Lấy danh sách các sinh viên
            students = list(student_data.keys())

            # Hiển thị cửa sổ tìm kiếm nâng cao và lấy ID của sinh viên được chọn
            search_dialog = AdvancedSearchDialog(students)
            if search_dialog.exec():
                selected_student = search_dialog.get_selected_student()
                if selected_student in student_data:
                    student_info = student_data[selected_student]
                    self.display_search_result(selected_student, student_info)
                    # Hiển thị cửa sổ chọn trường để cập nhật bằng lớp UpdateDialog
                    update_dialog = UpdateDialog(["StudentID", "Name", "Email", "Faculty", "Major", "Year"])
                    if update_dialog.exec():
                        selected_field, new_value = update_dialog.get_selected_field_and_value()

                        if selected_field != "Select Field":
                            # Cập nhật giá trị mới vào dữ liệu và cơ sở dữ liệu
                            student_info[selected_field] = new_value
                            students_ref.child(selected_student).set(student_info)
                            qmb_custom('Update Student', 'Student updated successfully.')
                            self.display_search_result(selected_student, student_info)
                        else:
                            qmb_custom('Update Student', 'Please select a valid field to update.')
                elif selected_student == "":
                    qmb_custom('Update Student', 'Please enter a Student ID.')
                    self.update_data()
                else:
                    qmb_custom('Update Student', 'Student ID not found.')
                    self.update_data()
        else:
            qmb_custom("Warning", "Please select a class.")
                
    def delete_data(self):
        if student_data:
            students = list(student_data.keys())
            advanced_search_dialog = AdvancedSearchDialog(students)
            # Hiển thị hộp thoại tìm kiếm nâng cao
            if advanced_search_dialog.exec():
                student_id = advanced_search_dialog.get_selected_student()
                if student_id:
                    student_info = student_data.get(student_id)
                    if student_info:
                        msgBox = QtWidgets.QMessageBox()
                        msgBox.setWindowTitle('Delete Student')
                        msgBox.setText('Are you sure you want to delete this student?')

                        # Thêm các nút và thiết lập kiểu nút
                        msgBox.addButton(QtWidgets.QMessageBox.StandardButton.Yes)
                        msgBox.addButton(QtWidgets.QMessageBox.StandardButton.No)
                        msgBox.setStyleSheet(
                            "QLabel{font-size: 20px; min-height:150 px; min-width: 400px;} QPushButton{ width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);}"
                        )
                        confirm = msgBox.exec()

                        if confirm == QtWidgets.QMessageBox.StandardButton.Yes:
                            # Xóa dữ liệu từ Firebase
                            students_ref.child(student_id).delete()
                            qmb_custom('Delete Student', 'Student deleted successfully.')

                            # Xóa dữ liệu từ student_data
                            del student_data[student_id]

                            # Cập nhật lại dữ liệu trên Table Widget
                            self.load_data()
                    else:
                        qmb_custom('Delete Student', 'Student ID not found.')
        else:
            qmb_custom("Warning", "Please select a class.")


    def calculate_data(self):
        if student_data:
            if self.tableWidget.rowCount() == 0:
                qmb_custom('Calculate', 'No data to calculate.')
                return

            # Lấy tên lớp đang được chọn
            selected_class_name = self.label_class.text().split(": ")[1]

            # Lấy thông tin về lớp được chọn từ dữ liệu lớp
            selected_class_info = class_data.get(selected_class_name)
            if not selected_class_info:
                qmb_custom('Calculate', 'Selected class information not found.')
                return

            # Lấy tổng số phiên học của lớp
            total_sessions = selected_class_info.get("TotalSessions", 0)
            if total_sessions == 0:
                qmb_custom('Calculate', 'Total sessions for the selected class is 0.')
                return

            # Tạo một danh sách các sinh viên thuộc lớp đã chọn
            selected_class_students = [student_id for student_id, student_info in student_data.items() if selected_class_name in student_info.get("Classes", [])]

            # Hiển thị hộp thoại cho người dùng để chọn liệu muốn sắp xếp điểm hay không
            msgBox = QtWidgets.QMessageBox()
            msgBox.setWindowTitle('Sort Points')
            msgBox.setText('Do you want to sort point?')

            # Thêm các nút tùy chỉnh
            msgBox.addButton(QtWidgets.QMessageBox.StandardButton.Yes)
            msgBox.addButton(QtWidgets.QMessageBox.StandardButton.No)

            msgBox.setStyleSheet(
                "QLabel{font-size: 20px; min-height:150 px; min-width: 400px;} "
                "QPushButton{ width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} "
                "QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);}"
            )

            reply = msgBox.exec()
            
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:  # Default sorting
                msgBox1 = QtWidgets.QMessageBox()
                msgBox1.setWindowTitle('Sort Points')
                msgBox1.setText('Choose type of sort\nYes = Ascending\nNo = Descending')

                # Thêm các nút tùy chỉnh
                msgBox1.addButton(QtWidgets.QMessageBox.StandardButton.Yes)
                msgBox1.addButton(QtWidgets.QMessageBox.StandardButton.No)

                msgBox1.setStyleSheet(
                    "QLabel{font-size: 20px; min-height:150 px; min-width: 400px;} "
                    "QPushButton{ width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} "
                    "QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);}"
                )
                reply1 = msgBox1.exec()
                if reply1 == QtWidgets.QMessageBox.StandardButton.Yes:
                    selected_class_students.sort(key=lambda student_id: student_data[student_id]['Classes'][selected_class_name].get('AttendanceCount', 0), reverse=False)
                else:
                    selected_class_students.sort(key=lambda student_id: student_data[student_id]['Classes'][selected_class_name].get('AttendanceCount', 0), reverse=True)
                
            else:
                reply = None

            for row, student_id in enumerate(selected_class_students):
                student_info = student_data.get(student_id)
                if student_info:
                    # Lấy giá trị AttendanceCount từ dữ liệu của sinh viên
                    attendance_count = student_info.get("Classes", {}).get(selected_class_name, {}).get("AttendanceCount", 0)

                    # Tính điểm dựa trên AttendanceCount chia cho TotalSessions
                    if total_sessions != 0:
                        points = attendance_count / total_sessions * 10
                    else:
                        points = 0
                    points_item = QtWidgets.QTableWidgetItem(str(round(points, 2)))  # làm tròn điểm đến 2 chữ số thập phân
                    self.tableWidget.setItem(row, 6, points_item)
                    set_read_only_flags(self.tableWidget)
        else:
            qmb_custom("Warning", "Please select a class.")

    

    def export_data(self):
        if student_data:
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Save File", "", "Excel Files (*.xlsx)")
            if file_path:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Student Data"

                # Ghi tiêu đề cột
                for column in range(self.tableWidget.columnCount()):
                    sheet.cell(row=1, column=column + 1, value=self.tableWidget.horizontalHeaderItem(column).text())

                # Ghi dữ liệu từ tableWidget vào file Excel
                for row in range(self.tableWidget.rowCount()):
                    for column in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, column)
                        if item is not None:
                            sheet.cell(row=row + 2, column=column + 1, value=item.text())

                try:
                    workbook.save(file_path)
                    qmb_custom('Export', f'Data exported to {file_path} successfully.')
                except Exception as e:
                    qmb_custom('Export', f'Error occurred while exporting data: {str(e)}')
        else:
            qmb_custom("Warning", "Please select a class.")
            
def set_read_only_flags(table_widget):
    for row in range(table_widget.rowCount()):
        for column in range(table_widget.columnCount()):
            item = table_widget.item(row, column)
            if item:
                item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)

def qmb_custom(string1, string2):
    msg_box = QtWidgets.QMessageBox()
    msg_box.setWindowTitle(string1)
    msg_box.setText(string2)
    # Thiết lập StyleSheet để căn giữa văn bản
    msg_box.setStyleSheet(
        "QLabel{font-size: 20px; min-height:150 px; min-width: 400px;} QPushButton{ width:100px; height:30px; border-radius: 5px; font-size: 15px; background-color: rgb(165, 213, 255);} QPushButton:hover{background-color: rgb(3, 105, 161); color: rgb(255,255,255);}"
        )
    msg_box.exec()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    StudentManagement = QtWidgets.QMainWindow()
    ui = Ui_StudentManagement()
    ui.setupUi(StudentManagement)
    StudentManagement.show()
    sys.exit(app.exec())